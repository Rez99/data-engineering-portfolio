from pathlib import Path
import zipfile
import tempfile

import pandas as pd
from openpyxl import load_workbook
import re


# ---------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------

INPUT_FILE = Path("money-manager-2.xlsx")
OUTPUT_FILE = Path("cell_table.csv")


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def open_workbook(path: Path):
    """
    Open an Excel workbook.
    """
    if path.suffix.lower() != ".xlsx":
        raise ValueError("Input file must be an .xlsx workbook.")

    return load_workbook(
        filename=path,
        data_only=False,  # preserve formulas
        read_only=True
    )


def parse_precedents(cell):
    """
    Extract cell references from an Excel formula.

    Examples:
        =A1+B1                  -> ["A1", "B1"]
        =SUM(B2:B10)            -> ["B2", "B10"]
        =IF(C2>0,D2,E2)         -> ["C2", "D2", "E2"]
        =Budget!D12             -> ["Budget!D12"]
        ='My Sheet'!F7 + A1     -> ["'My Sheet'!F7", "A1"]
    """
    if cell.data_type != "f":
        return None

    formula = str(cell.value)

    pattern = r"(?:'[^']+'|[A-Za-z0-9_]+)?!?[A-Z]{1,3}[0-9]+"

    matches = re.findall(pattern, formula)

    if not matches:
        return None

    # Remove duplicates while preserving order.
    matches = list(dict.fromkeys(matches))

    return ",".join(matches)

def build_dependents(df):
    """
    Given a cell table with columns:
        sheet, address, precedents

    return a list of dependents for each row.
    """

    # Create a unique identifier for every cell.
    cell_keys = df["sheet"] + "!" + df["address"]

    # Build map: precedent -> list of dependent cells.
    dependent_map = {}

    for i, row in df.iterrows():
        precedents = row["precedents"]

        if pd.isna(precedents) or precedents == "None":
            continue

        for precedent in str(precedents).split(","):
            precedent = precedent.strip()

            if not precedent:
                continue

            # Assume same sheet if not explicitly specified.
            if "!" not in precedent:
                precedent = f"{row['sheet']}!{precedent}"

            dependent_map.setdefault(precedent, []).append(cell_keys[i])

    # Build output column.
    dependents = []

    for key in cell_keys:
        refs = dependent_map.get(key, [])
        dependents.append(",".join(refs) if refs else None)

    return dependents

# ---------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------

def main():
    wb = open_workbook(INPUT_FILE)

    rows = []

    for ws in wb.worksheets:
        print(f"Processing sheet: {ws.title}")

        for row in ws.iter_rows():
            for cell in row:
                if not hasattr(cell, "coordinate"):
                    continue

                value = cell.value
                precedents = None

                if cell.data_type == "f":
                    precedents = parse_precedents(cell)

                rows.append(
                    {
                        "sheet": ws.title,
                        "address": cell.coordinate,
                        "value": value,
                        "precedents": precedents,
                        "dependents": None,
                    }
                )

    df = pd.DataFrame(
        rows,
        columns=[
            "sheet",
            "address",
            "value",
            "precedents",
            "dependents",
        ],
    )

    df["dependents"] = build_dependents(df)

    print()
    print(f"Rows: {len(df):,}")
    print(df.head(20))

    for col in ["value", "precedents", "dependents"]:
        df[col] = df[col].astype(str)

    df.to_csv(OUTPUT_FILE, index=False)

    print()
    print(f"Wrote: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()